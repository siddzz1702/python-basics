import os
import getpass
from cryptography.fernet import Fernet
from openpyxl import Workbook, load_workbook

FILE_NAME = "password_manager.xlsx"
KEY_FILE = "secret.key"

# ---------------- Security Setup ----------------
def load_key():
    if not os.path.exists(KEY_FILE):
        key = Fernet.generate_key()
        with open(KEY_FILE, "wb") as f:
            f.write(key)
    else:
        with open(KEY_FILE, "rb") as f:
            key = f.read()
    return Fernet(key)

fernet = load_key()

# ---------------- Excel Setup ----------------
def get_workbook():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        wb.save(FILE_NAME)
    return load_workbook(FILE_NAME)

def create_user(username, pin):
    wb = get_workbook()
    if username in wb.sheetnames:
        print("User already exists!")
        return
    ws = wb.create_sheet(username)
    ws.append(["Encrypted_PIN", fernet.encrypt(pin.encode()).decode()])
    ws.append(["App/Email", "Login ID", "Password (Encrypted)"])
    wb.save(FILE_NAME)
    print(f"User {username} created successfully!")

def verify_user(username, pin):
    wb = get_workbook()
    if username not in wb.sheetnames:
        print("No such user.")
        return False
    ws = wb[username]
    stored_pin = ws.cell(1, 2).value
    return fernet.decrypt(stored_pin.encode()).decode() == pin

# ---------------- Password Operations ----------------
def add_password(username, pin, app, login, password):
    if not verify_user(username, pin):
        print("Invalid PIN!")
        return
    wb = get_workbook()
    ws = wb[username]
    enc_password = fernet.encrypt(password.encode()).decode()
    ws.append([app, login, enc_password])
    wb.save(FILE_NAME)
    print("Password added successfully!")

def view_passwords(username, pin):
    if not verify_user(username, pin):
        print("Invalid PIN!")
        return
    wb = get_workbook()
    ws = wb[username]
    print(f"\nStored credentials for {username}:")
    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=1):
        app, login, enc_pass = row
        password = fernet.decrypt(enc_pass.encode()).decode()
        print(f"{idx}. App/Email: {app}, Login: {login}, Password: {password}")

def update_password(username, pin, entry_no, new_password):
    if not verify_user(username, pin):
        print("Invalid PIN!")
        return
    wb = get_workbook()
    ws = wb[username]
    row_no = entry_no + 2
    enc_password = fernet.encrypt(new_password.encode()).decode()
    ws.cell(row=row_no, column=3).value = enc_password
    wb.save(FILE_NAME)
    print("Password updated successfully!")

def delete_password(username, pin, entry_no):
    if not verify_user(username, pin):
        print("Invalid PIN!")
        return
    wb = get_workbook()
    ws = wb[username]
    row_no = entry_no + 2
    ws.delete_rows(row_no)
    wb.save(FILE_NAME)
    print("Password deleted successfully!")

# ---------------- Main Program ----------------
def main():
    while True:
        print("\n--- Password Manager ---")
        print("1. Create User")
        print("2. Add Password")
        print("3. View Passwords")
        print("4. Update Password")
        print("5. Delete Password")
        print("6. Exit")
        choice = input("Enter choice: ")

        if choice == "1":
            user = input("Enter username: ")
            pin = getpass.getpass("Set a PIN: ")
            create_user(user, pin)

        elif choice == "2":
            user = input("Enter username: ")
            pin = getpass.getpass("Enter PIN: ")
            app = input("Enter App/Email name: ")
            login = input("Enter Login ID: ")
            pwd = getpass.getpass("Enter Password: ")
            add_password(user, pin, app, login, pwd)

        elif choice == "3":
            user = input("Enter username: ")
            pin = getpass.getpass("Enter PIN: ")
            view_passwords(user, pin)

        elif choice == "4":
            user = input("Enter username: ")
            pin = getpass.getpass("Enter PIN: ")
            view_passwords(user, pin)
            entry_no = int(input("Enter entry number to update: "))
            new_pwd = getpass.getpass("Enter new password: ")
            update_password(user, pin, entry_no, new_pwd)

        elif choice == "5":
            user = input("Enter username: ")
            pin = getpass.getpass("Enter PIN: ")
            view_passwords(user, pin)
            entry_no = int(input("Enter entry number to delete: "))
            delete_password(user, pin, entry_no)

        elif choice == "6":
            print("Goodbye!")
            break
        else:
            print("Invalid choice, try again.")

if __name__ == "__main__":
    main()
